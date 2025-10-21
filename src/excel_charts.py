# src/excel_charts.py
"""
Adds:
1) 'Charts' sheet with visuals (Pie + Bar)
2) Enhancement data sheets:
   - Doors_Windows  (from data/output/doors_windows.json)
   - Flooring       (from data/output/flooring.json)
   - Area_Summary   (from data/output/area_summary.json)
"""

import sys, os, json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import PieChart, BarChart, Reference

# Enhancement JSON file paths
ENH_FILES = {
    "Doors_Windows": os.path.join("data", "output", "doors_windows.json"),
    "Flooring":      os.path.join("data", "output", "flooring.json"),
    "Area_Summary":  os.path.join("data", "output", "area_summary.json"),
}

def read_json(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def auto_fit(ws):
    """Auto-size Excel columns."""
    for col_cells in ws.columns:
        max_length = 0
        col = col_cells[0].column_letter
        for cell in col_cells:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col].width = max_length + 2

def _to_cell_value(v):
    """Convert Python objects to something Excel can store."""
    if isinstance(v, (dict, list)):
        try:
            return json.dumps(v, ensure_ascii=False, separators=(",", ":"))
        except Exception:
            return str(v)
    return v

def add_table(ws, data):
    """Write dict or list[dict] to the worksheet (safely)."""
    if isinstance(data, dict):
        r = 1
        for k, v in data.items():
            ws[f"A{r}"] = k
            ws[f"B{r}"] = _to_cell_value(v)
            ws[f"A{r}"].font = Font(bold=True)
            r += 1
    elif isinstance(data, list):
        if not data:
            ws["A1"] = "No data available"
            return
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append([_to_cell_value(row.get(h, "")) for h in headers])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    auto_fit(ws)

def integrate_enhancements(wb):
    """Add enhancement sheets if their JSONs exist."""
    for sheet_name, path in ENH_FILES.items():
        data = read_json(path)
        if not data:
            continue

        # Create or clear the sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws["A1:Z200"]:
                for c in row:
                    c.value = None
        else:
            ws = wb.create_sheet(sheet_name)

        # Write data
        if sheet_name == "Doors_Windows":
            ws.append(["Category", "Type", "Count", "Area m² (each)", "Rate per m²", "Amount"])
            for cat in ["doors", "windows"]:
                for it in data.get(cat, []):
                    ws.append([
                        cat.capitalize(),
                        it.get("type", ""),
                        it.get("count", 0),
                        it.get("area_m2_each", 0),
                        it.get("rate_per_m2", 0),
                        it.get("amount", 0),
                    ])
            ws.append([])
            totals = data.get("totals", {})
            ws.append(["", "", "", "", "Total", totals.get("total_amount", 0)])
            auto_fit(ws)

        elif sheet_name in ("Flooring", "Area_Summary"):
            add_table(ws, data)

def create_charts(wb):
    """Your original Charts logic, unchanged."""
    if "Summary" not in wb.sheetnames:
        sys.exit("[Error] 'Summary' sheet not found. Did you generate the workbook with rates_export.py?")

    # Remove old Charts sheet if present (so re-runs are clean)
    if "Charts" in wb.sheetnames:
        wb.remove(wb["Charts"])

    ws_sum = wb["Summary"]
    ws = wb.create_sheet("Charts")

    # Title
    ws["A1"] = "Visualization – Final Estimate"
    ws["A1"].font = Font(bold=True, size=14)

    # Grand Total callout (Summary!B12)
    ws["A3"] = "Grand Total"
    ws["A3"].font = Font(bold=True, size=12)
    ws["B3"] = ws_sum["B12"].value  # copy the number
    ws["B3"].number_format = ws_sum["B5"].number_format or "#,##0.00"
    ws["B3"].font = Font(bold=True, size=14)
    ws["B3"].alignment = Alignment(horizontal="left")

    # Pie: Materials vs Labor (Summary!A5:A6 and B5:B6)
    pie = PieChart()
    pie.title = "Materials vs Labor"
    data = Reference(ws_sum, min_col=2, min_row=5, max_row=6)      # B5:B6
    labels = Reference(ws_sum, min_col=1, min_row=5, max_row=6)    # A5:A6
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.height = 12
    pie.width = 18
    ws.add_chart(pie, "A5")

    # Bar: Overheads, Contingency, Profit, Tax (Summary!A8:A11 & B8:B11)
    bar = BarChart()
    bar.title = "Overheads / Contingency / Profit / Tax"
    bar.y_axis.title = "Amount"
    bar.x_axis.title = "Component"
    data2 = Reference(ws_sum, min_col=2, min_row=8, max_row=11)     # B8:B11
    cats2 = Reference(ws_sum, min_col=1, min_row=8, max_row=11)     # A8:A11
    bar.add_data(data2, titles_from_data=False)
    bar.set_categories(cats2)
    bar.height = 12
    bar.width = 24
    ws.add_chart(bar, "J5")

    # Neaten columns a bit
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20

    print("OK: Charts added to Excel.")

def main():
    xlsx_path = Path("data/output/final_estimate.xlsx")
    if not xlsx_path.exists():
        sys.exit("[Error] Excel not found. Run rates_export.py first to generate data/output/final_estimate.xlsx")

    wb = load_workbook(xlsx_path)

    # Step 1: add enhancement sheets (non-breaking)
    integrate_enhancements(wb)

    # Step 2: rebuild Charts sheet (your original)
    create_charts(wb)

    wb.save(xlsx_path)
    print(f"Updated workbook: {xlsx_path}")

if __name__ == "__main__":
    main()
