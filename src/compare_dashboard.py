# src/compare_dashboard.py
"""
Builds a comparative dashboard for India vs USA:
- Reads:
    data/output/qty_india_total.json
    data/output/qty_usa.json
    data/output/final_estimate.xlsx  (created earlier)
- Writes:
    - Adds/refreshes a 'Compare' sheet with a table and charts:
        * Clustered bar: Materials / Labor / Grand Total (IN vs US)
        * Percent-stacked bar: Materials vs Labor distribution by region
    - PNG image preview at data/output/compare_preview.png

Run:
  python src/compare_dashboard.py
"""

from pathlib import Path
import json
import sys

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.chart import BarChart, Reference

# For PNG preview (separate from Excel charts rendering)
import matplotlib.pyplot as plt


def load_json(p: Path):
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)


def safe_get(d, *keys, default=0.0):
    cur = d
    for k in keys:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    try:
        return float(cur)
    except Exception:
        return default


def ensure_wb(path_xlsx: Path):
    if path_xlsx.exists():
        return load_workbook(path_xlsx)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    wb.save(path_xlsx)
    return wb


def main():
    # Inputs
    in_path = Path("data/output/qty_india_total.json")
    us_path = Path("data/output/qty_usa.json")
    xlsx_path = Path("data/output/final_estimate.xlsx")
    png_path = Path("data/output/compare_preview.png")

    # Check inputs
    if not in_path.exists() and not us_path.exists():
        sys.exit("[Error] Neither India nor USA quantity JSON exists. Run earlier steps first.")

    india = load_json(in_path) if in_path.exists() else {}
    usa   = load_json(us_path) if us_path.exists() else {}

    # Pull the key totals
    IN_mat = safe_get(india, "totals", "materials_cost_subtotal", default=0.0)
    IN_lab = safe_get(india, "totals", "labor_cost_subtotal", default=0.0)
    IN_tot = safe_get(india, "totals", "grand_total", default=IN_mat + IN_lab)

    US_mat = safe_get(usa, "totals", "materials_cost_subtotal", default=0.0)
    US_lab = safe_get(usa, "totals", "labor_cost_subtotal", default=0.0)
    US_tot = safe_get(usa, "totals", "grand_total", default=US_mat + US_lab)

    # Open workbook (create if missing)
    wb = ensure_wb(xlsx_path)

    # Remove old Compare sheet
    if "Compare" in wb.sheetnames:
        wb.remove(wb["Compare"])

    ws = wb.create_sheet("Compare")

    # Title
    ws["A1"] = "India vs USA – Comparative Costs"
    ws["A1"].font = Font(bold=True, size=14)

    # Table headers
    ws["A3"] = "Category"; ws["B3"] = "INDIA"; ws["C3"] = "USA"
    for c in ("A3","B3","C3"):
        ws[c].font = Font(bold=True)

    # Rows: Materials, Labor, Grand Total
    rows = [
        ("Materials", IN_mat, IN_mat if False else IN_mat),  # placeholder replaced below
        ("Labor", IN_lab, US_lab),
        ("Grand Total", IN_tot, US_tot)
    ]
    rows[0] = ("Materials", IN_mat, US_mat)  # fix placeholder after vars exist

    r = 4
    for name, v_in, v_us in rows:
        ws[f"A{r}"] = name
        ws[f"B{r}"] = v_in
        ws[f"C{r}"] = v_us
        ws[f"B{r}"].number_format = "#,##0.00"
        ws[f"C{r}"].number_format = "#,##0.00"
        r += 1

    # Percent distribution table (materials vs labor only)
    ws["A8"] = "Distribution"; ws["A8"].font = Font(bold=True)
    ws["A9"] = "Component"; ws["B9"] = "INDIA"; ws["C9"] = "USA"
    for c in ("A9","B9","C9"):
        ws[c].font = Font(bold=True)
    ws["A10"] = "Materials"; ws["B10"] = IN_mat; ws["C10"] = US_mat
    ws["A11"] = "Labor";    ws["B11"] = IN_lab; ws["C11"] = US_lab
    ws["B10"].number_format = ws["C10"].number_format = "#,##0.00"
    ws["B11"].number_format = ws["C11"].number_format = "#,##0.00"

    # Clustered bar: Materials/Labor/Grand Total (IN vs US)
    bar = BarChart()
    bar.type = "col"
    bar.title = "IN vs US – Materials / Labor / Grand Total"
    bar.y_axis.title = "Amount"
    bar.x_axis.title = "Category"

    data = Reference(ws, min_col=2, min_row=3, max_col=3, max_row=6)  # include header row 3 and rows 4..6
    cats = Reference(ws, min_col=1, min_row=4, max_row=6)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.width = 28
    bar.height = 14
    ws.add_chart(bar, "E3")

    # Percent-stacked bar: Materials vs Labor distribution by region
    pbar = BarChart()
    pbar.type = "col"
    pbar.grouping = "percentStacked"
    pbar.title = "Materials vs Labor (%) – IN vs US"
    pbar.y_axis.title = "Percent"
    pbar.y_axis.scaling.max = 100
    pbar.y_axis.scaling.min = 0

    data2 = Reference(ws, min_col=2, min_row=9, max_col=3, max_row=11)  # header row 9 + rows 10..11
    cats2 = Reference(ws, min_col=1, min_row=10, max_row=11)
    pbar.add_data(data2, titles_from_data=True)
    pbar.set_categories(cats2)
    pbar.width = 28
    pbar.height = 14
    ws.add_chart(pbar, "E20")

    # Formats & widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    # Save workbook
    wb.save(xlsx_path)

    # -------- PNG preview with matplotlib (side-by-side bars for materials & labor) --------
    labels = ["Materials", "Labor"]
    IN_vals = [IN_mat, IN_lab]
    US_vals = [US_mat, US_lab]

    fig = plt.figure(figsize=(8, 4.5))
    x = range(len(labels))
    width = 0.35

    plt.bar([i - width/2 for i in x], IN_vals, width, label="India")
    plt.bar([i + width/2 for i in x], US_vals, width, label="USA")
    plt.xticks(list(x), labels)
    plt.ylabel("Amount")
    plt.title("India vs USA – Materials & Labor")
    plt.legend()
    png_path.parent.mkdir(parents=True, exist_ok=True)
    fig.tight_layout()
    fig.savefig(png_path, dpi=150)
    plt.close(fig)

    print("OK: Comparative dashboard updated.")
    print(f"Excel: {xlsx_path}  (sheet: Compare)")
    print(f"Preview PNG: {png_path}")

if __name__ == "__main__":
    main()
