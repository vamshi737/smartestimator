# src/validate_totals.py
from pathlib import Path
import json
import math
import openpyxl

JSON_PATH = Path("data/output/final_breakdown.json")
XLSX_PATH = Path("data/output/final_estimate.xlsx")

def find_value(ws, row_label: str, label_col="A", value_col="B"):
    """Find a row whose label cell equals row_label and return the value from value_col."""
    for r in range(1, ws.max_row + 1):
        lab = ws[f"{label_col}{r}"].value
        if isinstance(lab, str) and lab.strip().lower() == row_label.strip().lower():
            return ws[f"{value_col}{r}"].value
    return None

def pretty(v):
    try:
        return f"{float(v):,.2f}"
    except Exception:
        return str(v)

def close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False

def main():
    if not JSON_PATH.exists():
        print(f"[ERROR] JSON not found: {JSON_PATH}")
        return
    if not XLSX_PATH.exists():
        print(f"[ERROR] Excel not found: {XLSX_PATH}")
        return

    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    # ---- JSON values (structure seen in your file) ----
    subtotals = data.get("subtotals", {})
    j_in_mat = float(subtotals.get("india", {}).get("materials", 0) or 0)
    j_in_lab = float(subtotals.get("india", {}).get("labor", 0) or 0)
    j_us_mat = float(subtotals.get("usa", {}).get("materials", 0) or 0)
    j_us_lab = float(subtotals.get("usa", {}).get("labor", 0) or 0)

    j_sum = data.get("summary", {}) or {}
    j_mat_total = float(j_sum.get("materials", 0) or 0)
    j_lab_total = float(j_sum.get("labor", 0) or 0)
    j_grand = float(j_sum.get("grand_total", 0) or 0)

    # ---- Excel: Summary sheet (auto-find labels) ----
    ws_sum = wb["Summary"]
    x_mat_total = find_value(ws_sum, "Materials Subtotal")
    x_lab_total = find_value(ws_sum, "Labor Subtotal")
    x_grand = find_value(ws_sum, "Grand Total")

    # ---- Excel: Compare sheet (IN vs US materials/labor) ----
    # These are read for sanity; sheet layout may vary slightly, so we try to read the first table:
    try:
        ws_cmp = wb["Compare"]
        # Expected cells (from your screenshot):
        # B4 = India Materials, C4 = USA Materials, B5 = India Labor, C5 = USA Labor
        x_in_mat = ws_cmp["B4"].value
        x_us_mat = ws_cmp["C4"].value
        x_in_lab = ws_cmp["B5"].value
        x_us_lab = ws_cmp["C5"].value
    except Exception:
        x_in_mat = x_us_mat = x_in_lab = x_us_lab = None

    # ---- Print & Check ----
    print("== JSON (from final_breakdown.json) ==")
    print(f"India  Materials: {pretty(j_in_mat)}   Labor: {pretty(j_in_lab)}   (Sum: {pretty(j_in_mat + j_in_lab)})")
    print(f"USA    Materials: {pretty(j_us_mat)}   Labor: {pretty(j_us_lab)}   (Sum: {pretty(j_us_mat + j_us_lab)})")
    print(f"Totals Materials: {pretty(j_mat_total)} Labor: {pretty(j_lab_total)} Grand: {pretty(j_grand)}")

    print("\n== Excel (Summary sheet) ==")
    print(f"Materials Subtotal: {pretty(x_mat_total)}")
    print(f"Labor Subtotal    : {pretty(x_lab_total)}")
    print(f"Grand Total       : {pretty(x_grand)}")

    if x_in_mat is not None:
        print("\n== Excel (Compare sheet) ==")
        print(f"India  Materials: {pretty(x_in_mat)}   Labor: {pretty(x_in_lab)}")
        print(f"USA    Materials: {pretty(x_us_mat)}   Labor: {pretty(x_us_lab)}")

    # ---- Assertions with tolerance ----
    issues = []

    # Summary-level checks
    if not close(j_mat_total, x_mat_total): issues.append("Materials total mismatch (JSON vs Excel Summary).")
    if not close(j_lab_total, x_lab_total): issues.append("Labor total mismatch (JSON vs Excel Summary).")
    # Grand total may include overheads/contingency/profit/tax; JSON summary.grand_total should match Excel grand total
    if not close(j_grand, x_grand): issues.append("Grand total mismatch (JSON vs Excel Summary).")

    # Compare sheet sanity (optional)
    if x_in_mat is not None:
        if not close(j_in_mat, x_in_mat): issues.append("India materials mismatch (JSON vs Excel Compare).")
        if not close(j_in_lab, x_in_lab): issues.append("India labor mismatch (JSON vs Excel Compare).")
        if not close(j_us_mat, x_us_mat): issues.append("USA materials mismatch (JSON vs Excel Compare).")
        if not close(j_us_lab, x_us_lab): issues.append("USA labor mismatch (JSON vs Excel Compare).")

    print("\n== Result ==")
    if issues:
        for i in issues:
            print("•", i)
        print("Status:  MISMATCH — please review the above items.")
    else:
        print("Status:  All key totals match (within tolerance).")

if __name__ == "__main__":
    main()
