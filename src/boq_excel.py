# src/boq_excel.py
"""
Adds/updates a 'BOQ' sheet in data/output/final_estimate.xlsx with:
- BOQ – INDIA      (bricks, plaster, paint, steel, labor lines)
- BOQ – USA        (studs, sheathing, drywall, insulation, labor lines)
- BOQ – OPENINGS   (Doors & Windows from data/output/doors_windows.json)
- BOQ – FLOORING   (Flooring/Marble/Tiles from data/output/flooring.json)
- AREAS (Info)     (wall / openings / net wall / floor / gross areas)

Reads:
  - data/output/qty_india_total.json
  - data/output/qty_usa.json
  - data/output/doors_windows.json   (optional)
  - data/output/flooring.json        (optional)
  - data/output/area_summary.json    (optional, info only)
  - data/prices.json
  - data/output/final_estimate.xlsx (must exist)

Writes:
  - Updates data/output/final_estimate.xlsx with BOQ sheet (recreated each run)
"""

from pathlib import Path
import json, sys
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ---------- helpers ----------
def load_json(p: Path, default=None):
    if p.exists():
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)
    return default


def get_first(d: Dict[str, Any], candidates: List[List[str]], default=0.0):
    """Try multiple key-paths; return first numeric value found."""
    for path in candidates:
        cur = d
        ok = True
        for k in path:
            if isinstance(cur, dict) and k in cur:
                cur = cur[k]
            else:
                ok = False
                break
        if ok:
            try:
                return float(cur)
            except Exception:
                continue
    return float(default)


def fmt_currency(ws, cell, currency_hint="INR"):
    # Use USD style if currency looks like USD/$, else generic 2-decimal
    if currency_hint in ("USD", "$"):
        ws[cell].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    else:
        ws[cell].number_format = "#,##0.00"


def autosize(ws):
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            v = "" if c.value is None else str(c.value)
            widths[c.column] = max(widths.get(c.column, 0), len(v))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, w + 2), 48)


def box_style(ws, rng):
    thin = Side(style="thin", color="999999")
    for row in ws[rng]:
        for c in row:
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            c.alignment = Alignment(vertical="center")


# ---------- build BOQ rows (INDIA / USA) ----------
def build_india_boq(india: Dict[str, Any], rates_in: Dict[str, Any]) -> List[Tuple[str, str, float, float, float]]:
    """
    Returns rows: (Item, Unit, Qty, Rate, Amount)
    Uses defensive key lookups so it works with slightly different JSONs.
    """
    rows = []

    # Quantities (defensive candidates)
    bricks_nos = get_first(india, [["brickwork","bricks_count_with_wastage"],
                                   ["brickwork","bricks_count_without_wastage"],
                                   ["bricks_with_wastage"]], 0)
    cem_bags_bw = get_first(india, [["mortar_brickwork","cement_bags"]], 0)
    sand_m3_bw  = get_first(india, [["mortar_brickwork","sand_m3"]], 0)

    pl_area_m2  = get_first(india, [["plaster","area_m2"]], 0)
    pl_cem_bags = get_first(india, [["plaster","cement_bags"]], 0)
    pl_sand_m3  = get_first(india, [["plaster","sand_m3"]], 0)

    # Extras from india_extras if present
    paint_area_m2 = get_first(india, [["paint","area_m2"], ["extras","paint","area_m2"]], 0)
    steel_kg = get_first(india, [["steel","kg"], ["extras","steel_kg"]], 0)

    # Brickwork volume to derive labor m3 if needed
    brickwork_vol_m3 = get_first(india, [["derived","vol_brickwork_m3"]], 0)

    # Rates
    r_brick = float(rates_in.get("brick_per_piece", 0) or 0)
    r_cement_bag = float(rates_in.get("cement_bag_50kg", 0) or 0)
    r_sand_cum   = float(rates_in.get("sand_per_cum", 0) or 0)
    r_pl_sqm     = float(rates_in.get("plaster_per_sqm", 0) or 0)
    r_paint_liter= float(rates_in.get("paint_per_liter", 0) or 0)  # if we later compute liters
    r_steel_kg   = float(rates_in.get("steel_per_kg", 0) or 0)
    r_lab_bw_m3  = float(rates_in.get("labor_brickwork_per_m3", 0) or 0)
    r_lab_pl_sqm = float(rates_in.get("labor_plaster_per_m2", 0) or 0)
    r_lab_p_m2c  = float(rates_in.get("labor_paint_per_m2_per_coat", 0) or 0)

    # Default assumptions for paint if we want a rough amount (can be adjusted later)
    coats = 2.0
    coverage_m2_per_liter = 10.0
    liters_needed = (paint_area_m2 * coats) / coverage_m2_per_liter if coverage_m2_per_liter > 0 else 0

    rows.append(("Bricks", "Nos", bricks_nos, r_brick, bricks_nos * r_brick))
    rows.append(("Cement (Brickwork)", "Bag", cem_bags_bw, r_cement_bag, cem_bags_bw * r_cement_bag))
    rows.append(("Sand (Brickwork)", "m3", sand_m3_bw, r_sand_cum, sand_m3_bw * r_sand_cum))

    # Plaster surface (prefer per m2 rate)
    if r_pl_sqm > 0 and pl_area_m2 > 0:
        rows.append(("Plaster (Surface)", "m2", pl_area_m2, r_pl_sqm, pl_area_m2 * r_pl_sqm))
    else:
        rows.append(("Plaster Cement", "Bag", pl_cem_bags, r_cement_bag, pl_cem_bags * r_cement_bag))
        rows.append(("Plaster Sand", "m3", pl_sand_m3, r_sand_cum, pl_sand_m3 * r_sand_cum))

    # Paint (labor est.)
    paint_amount_labor = paint_area_m2 * r_lab_p_m2c * coats if (paint_area_m2>0 and r_lab_p_m2c>0) else 0
    rows.append(("Paint Area (labor est., 2 coats)", "m2", paint_area_m2, r_lab_p_m2c, paint_amount_labor))
    if r_paint_liter > 0 and liters_needed > 0:
        rows.append(("Paint (material est.)", "Liter", liters_needed, r_paint_liter, liters_needed * r_paint_liter))

    # Steel (rough)
    rows.append(("Steel (rough)", "kg", steel_kg, r_steel_kg, steel_kg * r_steel_kg))

    # Labor lines
    rows.append(("Labor – Brickwork", "m3", brickwork_vol_m3, r_lab_bw_m3, brickwork_vol_m3 * r_lab_bw_m3))
    if pl_area_m2 > 0:
        rows.append(("Labor – Plaster", "m2", pl_area_m2, r_lab_pl_sqm, pl_area_m2 * r_lab_pl_sqm))

    return rows


def build_usa_boq(usa: Dict[str, Any], rates_us: Dict[str, Any]) -> List[Tuple[str, str, float, float, float]]:
    """Returns rows: (Item, Unit, Qty, Rate, Amount) for USA framing style."""
    rows = []

    studs = get_first(usa, [["framing","studs_pcs"], ["studs_pcs"], ["studs"]], 0)
    plates = get_first(usa, [["framing","plates_pcs"], ["plates_pcs"], ["plates"]], 0)
    sheath_48 = get_first(usa, [["sheathing","sheets_4x8"], ["sheathing_sheets_4x8"]], 0)
    sheath_412 = get_first(usa, [["sheathing","sheets_4x12"], ["sheathing_sheets_4x12"]], 0)
    drywall_48 = get_first(usa, [["drywall","sheets_4x8"], ["drywall_sheets_4x8"]], 0)
    drywall_412 = get_first(usa, [["drywall","sheets_4x12"], ["drywall_sheets_4x12"]], 0)
    insul_packs = get_first(usa, [["insulation","packs"], ["insulation_packs"]], 0)

    # Rates
    r_2x4 = float(rates_us.get("2x4_stud_per_piece", 0) or 0)
    r_pl_24 = float(rates_us.get("2x4_plate_per_piece", 0) or 0)
    r_sh_48 = float(rates_us.get("sheathing_4x8_per_sheet", 0) or 0)
    r_sh_412= float(rates_us.get("sheathing_4x12_per_sheet", 0) or 0)
    r_dw_48 = float(rates_us.get("drywall_4x8_per_sheet", 0) or 0)
    r_dw_412= float(rates_us.get("drywall_4x12_per_sheet", 0) or 0)
    r_ins   = float(rates_us.get("insulation_pack", 0) or 0)

    # Labor
    r_lab_stud    = float(rates_us.get("labor_frame_per_stud", 0) or 0)
    r_lab_sheath  = float(rates_us.get("labor_sheath_per_sheet", 0) or 0)
    r_lab_drywall = float(rates_us.get("labor_drywall_per_sheet", 0) or 0)
    r_lab_insul   = float(rates_us.get("labor_insul_per_pack", 0) or 0)

    # Material rows
    rows.append(("2x4 Studs", "pcs", studs, r_2x4, studs * r_2x4))
    rows.append(("Plates (2x4)", "pcs", plates, r_pl_24, plates * r_pl_24))
    rows.append(("Sheathing 4x8", "sheet", sheath_48, r_sh_48, sheath_48 * r_sh_48))
    rows.append(("Sheathing 4x12", "sheet", sheath_412, r_sh_412, sheath_412 * r_sh_412))
    rows.append(("Drywall 4x8", "sheet", drywall_48, r_dw_48, drywall_48 * r_dw_48))
    rows.append(("Drywall 4x12", "sheet", drywall_412, r_dw_412, drywall_412 * r_dw_412))
    rows.append(("Insulation Packs", "pack", insul_packs, r_ins, insul_packs * r_ins))

    # Labor rows
    rows.append(("Labor – Framing (per stud)", "pcs", studs, r_lab_stud, studs * r_lab_stud))
    rows.append(("Labor – Sheathing (per sheet)", "sheet", sheath_48 + sheath_412, r_lab_sheath, (sheath_48 + sheath_412) * r_lab_sheath))
    rows.append(("Labor – Drywall (per sheet)", "sheet", drywall_48 + drywall_412, r_lab_drywall, (drywall_48 + drywall_412) * r_lab_drywall))
    rows.append(("Labor – Insulation (per pack)", "pack", insul_packs, r_lab_insul, insul_packs * r_lab_insul))

    return rows


# ---------- NEW: build BOQ rows (Openings + Flooring) ----------
def build_openings_boq(doors_windows: Dict[str, Any]) -> List[Tuple[str, str, float, float, float]]:
    """
    From data/output/doors_windows.json
    Each item already has 'area_m2_each', 'count', 'rate_per_m2', and 'amount'.
    Qty = count * area_m2_each (m2). Unit = m2.
    """
    rows: List[Tuple[str, str, float, float, float]] = []
    if not isinstance(doors_windows, dict):
        return rows

    for cat in ("doors", "windows"):
        for it in doors_windows.get(cat, []):
            area_each = float(it.get("area_m2_each", 0) or 0)
            cnt = float(it.get("count", 0) or 0)
            qty_m2 = area_each * cnt
            rate = float(it.get("rate_per_m2", 0) or 0)
            amt = float(it.get("amount", 0) or (qty_m2 * rate))
            label = f"{cat.capitalize()} – {it.get('type', '')}".strip()
            rows.append((label, "m2", qty_m2, rate, amt))

    # optional explicit total row is handled by add_table()
    return rows


def build_flooring_boq(flooring: Dict[str, Any]) -> List[Tuple[str, str, float, float, float]]:
    """
    From data/output/flooring.json
    Qty = total_area_m2_with_wastage (m2) if present, else area_m2.
    """
    rows: List[Tuple[str, str, float, float, float]] = []
    if not isinstance(flooring, dict):
        return rows

    qty = float(
        flooring.get("total_area_m2_with_wastage", flooring.get("area_m2", 0.0)) or 0.0
    )
    rate = float(flooring.get("rate_per_m2", 0) or 0)
    amt = float(flooring.get("amount", 0) or (qty * rate))
    material = str(flooring.get("material", "Flooring")).strip() or "Flooring"

    rows.append((f"{material}", "m2", qty, rate, amt))
    return rows


# ---------- writer helpers ----------
def add_table(ws, start_row, title, rows, currency_hint="INR"):
    """Writes a titled table and returns the next empty row after the table."""
    bold = Font(bold=True)
    r = start_row

    ws[f"A{r}"] = title
    ws[f"A{r}"].font = Font(bold=True, size=13)
    r += 1

    headers = ["Item", "Unit", "Quantity", "Rate", "Amount"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r, column=i, value=h).font = bold
    r += 1

    start_tbl = r
    for item, unit, qty, rate, amt in rows:
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=unit)
        ws.cell(row=r, column=3, value=qty)
        ws.cell(row=r, column=4, value=rate)
        ws.cell(row=r, column=5, value=amt)
        # formats
        ws[f"C{r}"].number_format = "#,##0.00"
        fmt_currency(ws, f"D{r}", currency_hint)
        fmt_currency(ws, f"E{r}", currency_hint)
        r += 1

    end_tbl = r - 1

    # Add a Total row (works even if there were 0 data rows - then SUM over empty range = 0)
    ws.cell(row=r, column=4, value="Total").font = bold
    ws.cell(row=r, column=5, value=f"=SUM(E{start_tbl}:E{end_tbl})").font = bold
    fmt_currency(ws, f"E{r}", currency_hint)
    r += 2  # blank line after table

    # Include header row in box
    box_style(ws, f"A{start_tbl-1}:E{end_tbl}")
    return r


def add_info_table(ws, start_row, title, pairs):
    """Non-monetary info table (key/value). Returns next row."""
    bold = Font(bold=True)
    r = start_row
    ws[f"A{r}"] = title
    ws[f"A{r}"].font = Font(bold=True, size=13)
    r += 1

    ws.cell(row=r, column=1, value="Metric").font = bold
    ws.cell(row=r, column=2, value="Value").font = bold
    r += 1

    start_tbl = r
    for k, v in pairs:
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
        r += 1
    end_tbl = r - 1

    box_style(ws, f"A{start_tbl-1}:B{end_tbl}")
    r += 2
    return r


# ---------- main ----------
def main():
    base = Path("data/output")
    xlsx = base / "final_estimate.xlsx"
    india_json = base / "qty_india_total.json"
    usa_json   = base / "qty_usa.json"
    doors_windows_json = base / "doors_windows.json"
    flooring_json      = base / "flooring.json"
    area_summary_json  = base / "area_summary.json"
    prices_json = Path("data/prices.json")

    if not xlsx.exists():
        sys.exit("[Error] Excel not found: data/output/final_estimate.xlsx. Run earlier steps first.")

    # Load inputs
    india = load_json(india_json, {}) or {}
    usa   = load_json(usa_json, {}) or {}
    doors_windows = load_json(doors_windows_json, {}) or {}
    flooring = load_json(flooring_json, {}) or {}
    area_summary = load_json(area_summary_json, {}) or {}
    prices = load_json(prices_json, {}) or {}

    rates_in = prices.get("IN", {}) if isinstance(prices, dict) else {}
    rates_us = prices.get("US", {}) if isinstance(prices, dict) else {}
    currency_global = ""
    if isinstance(prices, dict) and "GLOBAL" in prices and isinstance(prices["GLOBAL"], dict):
        currency_global = str(prices["GLOBAL"].get("currency", "") or "")

    # Open workbook
    wb = load_workbook(xlsx)

    # Remove old BOQ if exists (clean re-run)
    if "BOQ" in wb.sheetnames:
        wb.remove(wb["BOQ"])
    ws = wb.create_sheet("BOQ")

    ws["A1"] = "Bill of Quantities (BOQ)"
    ws["A1"].font = Font(bold=True, size=14)

    # Currency hints
    in_currency_hint = currency_global if currency_global else "INR"
    us_currency_hint = "USD"  # display USD style for USA section

    r = 3

    # INDIA
    india_rows = build_india_boq(india, rates_in)
    r = add_table(ws, r, "BOQ – INDIA", india_rows, currency_hint=in_currency_hint)

    # USA
    usa_rows = build_usa_boq(usa, rates_us)
    r = add_table(ws, r, "BOQ – USA", usa_rows, currency_hint=us_currency_hint)

    # OPENINGS (Doors & Windows)
    open_rows = build_openings_boq(doors_windows)
    if open_rows:
        r = add_table(ws, r, "BOQ – OPENINGS (Doors & Windows)", open_rows,
                      currency_hint=in_currency_hint if currency_global else "USD")

    # FLOORING
    floor_rows = build_flooring_boq(flooring)
    if floor_rows:
        r = add_table(ws, r, "BOQ – FLOORING", floor_rows,
                      currency_hint=in_currency_hint if currency_global else "USD")

    # AREAS (Info only)
    if isinstance(area_summary, dict) and area_summary:
        pairs = [
            ("Wall Area (m²)",        float(area_summary.get("wall_area_m2", 0) or 0)),
            ("Openings Area (m²)",    float(area_summary.get("openings_area_m2", 0) or 0)),
            ("Net Wall Area (m²)",    float(area_summary.get("net_wall_area_m2", 0) or 0)),
            ("Floor Area (m²)",       float(area_summary.get("floor_area_m2", 0) or 0)),
            ("Gross Area (m²)",       float(area_summary.get("gross_area_m2", 0) or 0)),
        ]
        r = add_info_table(ws, r, "AREAS (Info)", pairs)

    autosize(ws)
    wb.save(xlsx)
    print("OK: BOQ sheet added/updated in", xlsx)

if __name__ == "__main__":
    main()
