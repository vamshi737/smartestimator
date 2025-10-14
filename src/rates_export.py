# src/rates_export.py
"""
Reads prices.json and one or both quantity JSONs (India total, USA),
applies overhead/contingency/profit/tax, and exports:
- data/output/final_breakdown.json
- data/output/final_estimate.xlsx
- data/output/final_estimate.pdf

Dependencies:
  pip install openpyxl reportlab
"""

from pathlib import Path
import json, argparse, sys
from typing import Dict, Any

# Excel
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, numbers
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet

def load_json(p: Path):
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)

def try_load(p: Path):
    return load_json(p) if p.exists() else None

def get_global(prices: Dict[str, Any], key: str, default=0.0):
    g = prices.get("GLOBAL", {})
    if isinstance(g, dict) and key in g:
        return g[key]
    return default

def to_float(x, default=0.0):
    try:
        return float(x)
    except Exception:
        return default

def safe_num(d: Dict[str, Any], *keys, default=0.0):
    cur = d
    for k in keys:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return to_float(cur, default)

def autosize_columns(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            v = str(cell.value) if cell.value is not None else ""
            widths[cell.column] = max(widths.get(cell.column, 0), len(v))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, w + 2), 40)

def add_table_style(ws, cell_range):
    thin = Side(border_style="thin", color="999999")
    rows = ws[cell_range]
    for r in rows:
        for c in r:
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            c.alignment = Alignment(vertical="center")

def build_excel(path_xlsx: Path, currency: str, summary: Dict[str, float], details: Dict[str, Any], prices: Dict[str, Any]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    bold = Font(bold=True)

    ws["A1"] = "Final Estimate Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Currency: {currency}"

    rows = [
        ["Materials Subtotal", summary["materials"]],
        ["Labor Subtotal", summary["labor"]],
        ["Subtotal (Materials + Labor)", summary["materials"] + summary["labor"]],
        ["Overheads", summary["overheads"]],
        ["Contingency", summary["contingency"]],
        ["Profit", summary["profit"]],
        ["Tax", summary["tax"]],
        ["Grand Total", summary["grand_total"]],
    ]

    ws.append([])
    ws.append(["Section", "Amount"])
    ws["A4"].font = bold; ws["B4"].font = bold

    r0 = 5
    for r in rows:
        ws.append(r)
    r1 = r0 + len(rows) - 1
    add_table_style(ws, f"A4:B{r1}")

    # Currency format
    for r in range(5, r1 + 1):
        ws[f"B{r}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE if currency in ("USD","$") else "#,##0.00"

    autosize_columns(ws)

    # Rates sheet (show GLOBAL if present)
    ws2 = wb.create_sheet("Rates")
    ws2["A1"] = "Pricing Knobs"
    ws2["A1"].font = Font(bold=True, size=12)
    g = prices.get("GLOBAL", {})
    ws2.append(["Key", "Value"])
    ws2["A3"].font = bold; ws2["B3"].font = bold
    r = 4
    for k in ["currency","overhead_pct","contingency_pct","profit_pct","tax_pct"]:
        ws2.append([k, g.get(k, "")])
        r += 1
    add_table_style(ws2, f"A3:B{r-1}")
    autosize_columns(ws2)

    # Details sheet: show inputs taken from India/USA files
    ws3 = wb.create_sheet("Details")
    ws3["A1"] = "Input Subtotals"
    ws3["A1"].font = Font(bold=True, size=12)
    ws3.append(["Source", "Materials", "Labor"])
    ws3["A3"].font = bold; ws3["B3"].font = bold; ws3["C3"].font = bold
    r = 4
    for src in ["india", "usa"]:
        v = details.get(src, {})
        ws3.append([src.upper(), v.get("materials", 0.0), v.get("labor", 0.0)])
        r += 1
    add_table_style(ws3, f"A3:C{r-1}")
    for c in range(4, r):
        ws3[f"B{c}"].number_format = "#,##0.00"
        ws3[f"C{c}"].number_format = "#,##0.00"
    autosize_columns(ws3)

    wb.save(path_xlsx)

def build_pdf(path_pdf: Path, currency: str, summary: Dict[str, float], details: Dict[str, Any], prices: Dict[str, Any]):
    doc = SimpleDocTemplate(str(path_pdf), pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Final Estimate Summary", styles["Title"]))
    story.append(Paragraph(f"Currency: {currency}", styles["Normal"]))
    story.append(Spacer(1, 6))

    data = [
        ["Section", "Amount"],
        ["Materials Subtotal", f"{summary['materials']:.2f}"],
        ["Labor Subtotal", f"{summary['labor']:.2f}"],
        ["Subtotal (Materials + Labor)", f"{(summary['materials']+summary['labor']):.2f}"],
        ["Overheads", f"{summary['overheads']:.2f}"],
        ["Contingency", f"{summary['contingency']:.2f}"],
        ["Profit", f"{summary['profit']:.2f}"],
        ["Tax", f"{summary['tax']:.2f}"],
        ["Grand Total", f"{summary['grand_total']:.2f}"],
    ]
    tbl = Table(data, hAlign="LEFT", colWidths=[80*mm, 50*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ("ALIGN", (1,1), (-1,-1), "RIGHT"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.white]),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 10))

    g = prices.get("GLOBAL", {})
    assumptions = [
        ["Assumption", "Value"],
        ["Overhead %", str(g.get("overhead_pct", 0))],
        ["Contingency %", str(g.get("contingency_pct", 0))],
        ["Profit %", str(g.get("profit_pct", 0))],
        ["Tax %", str(g.get("tax_pct", 0))],
    ]
    tbl2 = Table(assumptions, hAlign="LEFT", colWidths=[60*mm, 70*mm])
    tbl2.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ("ALIGN", (1,1), (-1,-1), "RIGHT"),
    ]))
    story.append(Paragraph("Assumptions", styles["Heading2"]))
    story.append(tbl2)

    doc.build(story)

def main():
    ap = argparse.ArgumentParser(description="Rates loader + Excel/PDF exports")
    ap.add_argument("--prices", required=True, help="data/prices.json")
    ap.add_argument("--in_json", default="data/output/qty_india_total.json", help="India merged quantities JSON (optional)")
    ap.add_argument("--us_json", default="data/output/qty_usa.json", help="USA quantities JSON (optional)")
    ap.add_argument("--currency", default="", help="Override currency code/symbol (optional)")
    ap.add_argument("--out_xlsx", default="data/output/final_estimate.xlsx")
    ap.add_argument("--out_pdf",  default="data/output/final_estimate.pdf")
    ap.add_argument("--out_json", default="data/output/final_breakdown.json")
    args = ap.parse_args()

    prices_path = Path(args.prices)
    if not prices_path.exists():
        sys.exit(f"[Error] prices file not found: {args.prices}")
    prices = load_json(prices_path)

    # Load available quantity files
    india_path = Path(args.in_json)
    usa_path   = Path(args.us_json)
    india = try_load(india_path)
    usa   = try_load(usa_path)

    if not india and not usa:
        sys.exit("[Error] No quantities found. Provide at least one of --in_json or --us_json that exists.")

    # Extract subtotals from each source
    details = {"india": {"materials": 0.0, "labor": 0.0}, "usa": {"materials": 0.0, "labor": 0.0}}
    if india:
        details["india"]["materials"] = safe_num(india, "totals", "materials_cost_subtotal", default=0.0)
        details["india"]["labor"]     = safe_num(india, "totals", "labor_cost_subtotal", default=0.0)
    if usa:
        details["usa"]["materials"] = safe_num(usa, "totals", "materials_cost_subtotal", default=0.0)
        details["usa"]["labor"]     = safe_num(usa, "totals", "labor_cost_subtotal", default=0.0)

    materials_sub = details["india"]["materials"] + details["usa"]["materials"]
    labor_sub     = details["india"]["labor"] + details["usa"]["labor"]
    subtotal      = materials_sub + labor_sub

    # Global multipliers
    currency = args.currency if args.currency else str(get_global(prices, "currency", ""))
    overhead_pct    = to_float(get_global(prices, "overhead_pct",    0.0))
    contingency_pct = to_float(get_global(prices, "contingency_pct", 0.0))
    profit_pct      = to_float(get_global(prices, "profit_pct",      0.0))
    tax_pct         = to_float(get_global(prices, "tax_pct",         0.0))

    overheads   = subtotal * (overhead_pct / 100.0)
    contingency = (subtotal + overheads) * (contingency_pct / 100.0)
    profit      = (subtotal + overheads + contingency) * (profit_pct / 100.0)
    tax         = (subtotal + overheads + contingency + profit) * (tax_pct / 100.0)
    grand_total = subtotal + overheads + contingency + profit + tax

    summary = {
        "materials": materials_sub,
        "labor": labor_sub,
        "overheads": overheads,
        "contingency": contingency,
        "profit": profit,
        "tax": tax,
        "grand_total": grand_total,
        "overhead_pct": overhead_pct,
        "contingency_pct": contingency_pct,
        "profit_pct": profit_pct,
        "tax_pct": tax_pct,
        "currency": currency
    }

    # Save breakdown JSON
    out_json = Path(args.out_json)
    out_json.parent.mkdir(parents=True, exist_ok=True)
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump({
            "inputs": {
                "prices_path": str(prices_path),
                "india_path": str(india_path) if india else "",
                "usa_path": str(usa_path) if usa else "",
                "currency": currency
            },
            "subtotals": details,
            "summary": summary
        }, f, indent=2)

    # Excel
    build_excel(Path(args.out_xlsx), currency or "CUR", summary, details, prices)
    # PDF
    build_pdf(Path(args.out_pdf), currency or "CUR", summary, details, prices)

    # Console summary
    print("OK: Rates applied and exports generated.")
    print(f"JSON: {args.out_json}")
    print(f"XLSX: {args.out_xlsx}")
    print(f"PDF : {args.out_pdf}")
    print("Totals:")
    print(f"  Materials:  {materials_sub:.2f}")
    print(f"  Labor:      {labor_sub:.2f}")
    print(f"  Overheads:  {overheads:.2f}  (pct={overhead_pct})")
    print(f"  Contingency:{contingency:.2f}  (pct={contingency_pct})")
    print(f"  Profit:     {profit:.2f}  (pct={profit_pct})")
    print(f"  Tax:        {tax:.2f}  (pct={tax_pct})")
    print(f"  Grand Total:{grand_total:.2f}")

if __name__ == "__main__":
    main()
